import openpyxl
import datetime
import get
import rewriter
from openpyxl.utils import range_boundaries
from openpyxl.workbook import Workbook
from get import file_names
from dateutil import relativedelta

def get_week_type(day): # Возвращает тип недели. True - числитель, False - знаменатель
    date = datetime.datetime.now()
    date = date + relativedelta.relativedelta(weekday=day)
    if date.isocalendar()[1] % 2 != 0:
        return True
    else:
        return False

def letter(string):
    string = string.replace('26', 'Z')
    string = string.replace('25', 'Y')
    string = string.replace('24', 'X')
    string = string.replace('23', 'W')
    string = string.replace('22', 'V')
    string = string.replace('21', 'U')
    string = string.replace('20', 'T')
    string = string.replace('19', 'S')
    string = string.replace('18', 'R')
    string = string.replace('17', 'Q')
    string = string.replace('16', 'P')
    string = string.replace('15', 'O')
    string = string.replace('14', 'N')
    string = string.replace('13', 'M')
    string = string.replace('12', 'L')
    string = string.replace('11', 'K')
    string = string.replace('10', 'J')
    string = string.replace('9', 'I')
    string = string.replace('8', 'H')
    string = string.replace('7', 'G')
    string = string.replace('6', 'F')
    string = string.replace('5', 'E')
    string = string.replace('4', 'D')
    string = string.replace('3', 'C')
    string = string.replace('2', 'B')
    string = string.replace('1', 'A')
    return string

def number_to_letters(number):
    #26
    base = 26
    newNum = ''
    while number > 0:
        newNum = letter(str(number % base)) + newNum
        number = number // base
    return newNum

def week_number_to_str(number):
    days_of_week = {0:'понедельник', 1:'вторник', 2:'среда', 3:'четверг', 4:'пятница', 5:'суббота', 6:'восскресение'}
    try:
        return days_of_week[number]
    except:
        return week_number_to_str(number % 7)

def week_str_to_number(string):
    days_of_week = {'понедельник':0, 'вторник':1, 'среда':2, 'четверг':3, 'пятница':4, 'суббота':5, 'восскресение':6, 'неделя':[], 'след_неделя':[]}
    try:
        return days_of_week[string]
    except:
        return False

def week_type_to_str(bool):
    date = datetime.datetime.now().weekday()
    string = ''
    if bool != get_week_type(date):
        string += 'следующей недели'
    else:
        string += 'текущей недели'

    if bool:
        string += ' [числитель]'
    else:
        string += ' [знаменатель]'

    return string

def format_sheet(sheet):
    counter = 0
    while len(sheet.merged_cells.ranges) > 0:
        counter += 1
        for cell_group in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            sheet.unmerge_cells(str(cell_group))
            for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value
    return counter

def delete_spaces(string):
    while string.find('  ') != -1:
        string = string.replace('  ', ' ')
    return string

def find_timetable(week_number, week_type, cells, coordinates, subgroup):
    output = ''
    title = f'{week_number_to_str(week_number).capitalize()} {week_type_to_str(week_type)}\n'
    if len(coordinates) == 2:
        for p in range(1, 8):
            if cells[ p * 2 - int(week_type) + week_number * 14 ][ coordinates[1][1] ].value != '' and cells[ p * 2 - int(week_type) + week_number * 14 ][ coordinates[1][1] ].value != None and subgroup == 2:
                output += f'| {rewriter.underline(cells[ p * 2 - int(week_type) + week_number * 14 ][1].value)}| '
                output += delete_spaces(f'{cells[ p * 2 - int(week_type) + week_number * 14 ][ coordinates[1][1] ].value}'.replace('\n', ' ') + '\n')
            elif cells[ p * 2 - int(week_type) + week_number * 14 ][ coordinates[0][1] ].value != '' and cells[ p * 2 - int(week_type) + week_number * 14 ][ coordinates[0][1] ].value != None and subgroup == 1:
                output += f'| {rewriter.underline(cells[ p * 2 - int(week_type) + week_number * 14 ][1].value)}| '
                output += delete_spaces(f'{cells[ p * 2 - int(week_type) + week_number * 14 ][ coordinates[0][1] ].value}'.replace('\n', ' ') + '\n')
    else:
        for p in range(1, 8):
            if cells[ p * 2 - int(week_type) + week_number * 14 ][ coordinates[0][1] ].value != '' and cells[ p * 2 - int(week_type) + week_number * 14 ][ coordinates[0][1] ].value != None:
                output += f'| {rewriter.underline(cells[ p * 2 - int(week_type) + week_number * 14 ][1].value)}| '
                output += delete_spaces(f'{cells[ p * 2 - int(week_type) + week_number * 14 ][ coordinates[0][1] ].value}'.replace('\n', ' ') + '\n')
    if output:
        return f'{title}{output}'
    else:
        return f'Нет ни одного занятия, назначенного на этот день ({week_number_to_str(week_number)})\n'

def find_coordinates(cells, week_type, group):
    coordinates = []
    for i in range(0, len(cells)):
        for j in range(0, len(cells[i])):
            try:
                if cells[i][j].value.lower() == group.lower():
                    if week_type:
                        coordinates.append([i + 1, j])
                    else:
                        coordinates.append([i + 2, j])
            except:
                continue
    return coordinates

def initialization():
    global file_names
    print('Инициализация..\n')
    for file_name in file_names.keys():
        book = openpyxl.load_workbook(file_names[file_name])
        for sheet_name in book.sheetnames:
            sheet = book[sheet_name]
            if format_sheet(sheet) > 0:
                book.save(file_names[file_name])
    print('Инициализация прошла успешно!')

def timetable_controller(group, subgroup, time, faculty):
    global file_names
    week_type = get_week_type(week_str_to_number(time))
    book_name = file_names[faculty]
    book = openpyxl.load_workbook(book_name)
    coordinates = []
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        if format_sheet(sheet) > 0:
            book.save(book_name)

        index = 1 # Если индекс нечетный, то числитель - всегда четная строка, а иначе числитель всегда нечетный
        while sheet[index][0].value != 'Дни' and index <= sheet.max_row:
            index += 1

        cells = list(sheet['A' + str(index):number_to_letters(sheet.max_column-1) + str(sheet.max_row)])

        coordinates = coordinates.clear()
        coordinates = find_coordinates(cells, week_type, group).copy()
        if len(coordinates) == 0:
            continue

        output = ""
        if subgroup == 1:
            output +="[1 подгруппа]\n"
        elif subgroup == 2:
            output +='[2 подгруппа]\n'

        if isinstance(week_str_to_number(time), list):
            if time == 'неделя':
                for i in range(0, 6):
                    output += find_timetable(i, week_type, cells, coordinates, subgroup)
                    output += '\n'
            elif time == 'след_неделя':
                for i in range(0, 6):
                    output += find_timetable(i, not week_type, cells, coordinates, subgroup)
                    output += '\n'
            else:
                print('Неверно указан временной промежуток (timetable_controller)')
                for i in range(0, 6):
                    output += find_timetable(i, week_type, cells, coordinates, subgroup)
                    output += '\n'
        else:
            output += find_timetable(week_str_to_number(time), week_type, cells, coordinates, subgroup)
        return output

initialization()
